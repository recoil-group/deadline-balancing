"""Print a compact, read-only view of an Excel workbook for coding agents.

The output favors workbook meaning over visual fidelity: populated cells, formula
results and patterns, sheet structure, comments, and hyperlinks. It never saves
the workbook or recalculates formulas.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable


DEFAULT_MAX_CELLS = 1200
DEFAULT_MAX_FORMULA_PATTERNS = 40
THREADED_COMMENT_MARKER = "[Threaded comment]"


@dataclass(frozen=True)
class FormulaInfo:
    coordinate: str
    column: str
    row: int
    expression: str
    cached_value: Any

    @property
    def has_cached_value(self) -> bool:
        return self.cached_value is not None


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Print a compact structural and textual view of an XLSX workbook."
    )
    parser.add_argument("workbook", type=Path, help="Workbook to inspect (.xlsx or .xlsm).")
    parser.add_argument(
        "--sheet",
        action="append",
        help="Inspect only this worksheet. Repeat to select multiple worksheets.",
    )
    parser.add_argument(
        "--range",
        dest="cell_range",
        help="Inspect one A1 range, such as A1:F20. Requires exactly one selected sheet.",
    )
    parser.add_argument(
        "--max-cells",
        type=int,
        default=DEFAULT_MAX_CELLS,
        help=(
            "Maximum populated cells printed per worksheet before truncation "
            f"(default: {DEFAULT_MAX_CELLS}; 0 means unlimited)."
        ),
    )
    parser.add_argument(
        "--show-formulas",
        action="store_true",
        help="Show every formula expression inline as well as the compact pattern summary.",
    )
    parser.add_argument(
        "--max-formula-patterns",
        type=int,
        default=DEFAULT_MAX_FORMULA_PATTERNS,
        help=(
            "Maximum formula patterns printed per worksheet "
            f"(default: {DEFAULT_MAX_FORMULA_PATTERNS}; 0 means unlimited)."
        ),
    )
    return parser.parse_args()


def format_value(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float):
        # Cached formula results often contain binary floating-point noise that
        # obscures the useful value for a reader (for example 3.77999999999999).
        rounded = round(value, 8)
        if rounded == 0:
            rounded = 0.0
        return format(rounded, ".12g")
    if isinstance(value, str):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def human_size(size: int) -> str:
    value = float(size)
    for unit in ("B", "KiB", "MiB", "GiB"):
        if value < 1024 or unit == "GiB":
            precision = 0 if unit == "B" else 1
            return f"{value:.{precision}f} {unit}"
        value /= 1024
    raise AssertionError("unreachable")


def actual_content_range(cells: list[Any]) -> str:
    from openpyxl.utils import get_column_letter

    if not cells:
        return "empty"
    first_row = min(cell.row for cell in cells)
    last_row = max(cell.row for cell in cells)
    first_column = min(cell.column for cell in cells)
    last_column = max(cell.column for cell in cells)
    return (
        f"{get_column_letter(first_column)}{first_row}:"
        f"{get_column_letter(last_column)}{last_row}"
    )


def selected_bounds(worksheet: Any, cell_range: str | None) -> tuple[int, int, int, int]:
    from openpyxl.utils.cell import range_boundaries

    if cell_range:
        try:
            first_column, first_row, last_column, last_row = range_boundaries(cell_range)
        except ValueError as error:
            raise ValueError(f"Invalid A1 range {cell_range!r}.") from error
        if not all((first_column, first_row, last_column, last_row)):
            raise ValueError(f"Range must be rectangular: {cell_range!r}.")
        return first_row, last_row, first_column, last_column

    first_column, first_row, last_column, last_row = range_boundaries(
        worksheet.calculate_dimension()
    )
    return first_row, last_row, first_column, last_column


def iter_selected_cells(worksheet: Any, bounds: tuple[int, int, int, int]) -> Iterable[Any]:
    first_row, last_row, first_column, last_column = bounds
    for row in worksheet.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=first_column,
        max_col=last_column,
    ):
        yield from row


def normalize_formula(expression: str, row: int) -> str:
    """Replace same-row references so copied formulas collapse into one pattern."""

    # Absolute row references (for example ``$A$13``) deliberately stay exact.
    same_row_reference = re.compile(
        rf"(?<![A-Za-z0-9_])(\$?[A-Z]{{1,3}}){row}(?!\d)"
    )
    return same_row_reference.sub(r"\1{row}", expression)


def contiguous_coordinate_ranges(formulas: list[FormulaInfo]) -> str:
    if not formulas:
        return ""

    by_column: dict[str, list[int]] = defaultdict(list)
    for formula in formulas:
        by_column[formula.column].append(formula.row)

    ranges: list[str] = []
    for column, rows in sorted(by_column.items()):
        ordered_rows = sorted(set(rows))
        start = previous = ordered_rows[0]
        for current in ordered_rows[1:]:
            if current == previous + 1:
                previous = current
                continue
            ranges.append(coordinate_range(column, start, previous))
            start = previous = current
        ranges.append(coordinate_range(column, start, previous))
    return ", ".join(ranges)


def coordinate_range(column: str, start: int, end: int) -> str:
    return f"{column}{start}" if start == end else f"{column}{start}:{column}{end}"


def compressed_numbers(numbers: Iterable[int]) -> str:
    ordered = sorted(set(numbers))
    if not ordered:
        return ""
    ranges: list[str] = []
    start = previous = ordered[0]
    for current in ordered[1:]:
        if current == previous + 1:
            previous = current
            continue
        ranges.append(str(start) if start == previous else f"{start}-{previous}")
        start = previous = current
    ranges.append(str(start) if start == previous else f"{start}-{previous}")
    return ", ".join(ranges)


def hidden_summary(worksheet: Any) -> str | None:
    from openpyxl.utils import column_index_from_string

    hidden_rows = [index for index, dimension in worksheet.row_dimensions.items() if dimension.hidden]
    hidden_columns: list[int] = []
    for key, dimension in worksheet.column_dimensions.items():
        if not dimension.hidden:
            continue
        start = dimension.min or column_index_from_string(key)
        end = dimension.max or start
        hidden_columns.extend(range(start, end + 1))

    parts = []
    if hidden_rows:
        parts.append(f"rows {compressed_numbers(hidden_rows)}")
    if hidden_columns:
        from openpyxl.utils import get_column_letter

        labels = [get_column_letter(index) for index in sorted(set(hidden_columns))]
        parts.append(f"columns {', '.join(labels)}")
    return "; ".join(parts) if parts else None


def clean_comment(text: str) -> tuple[str, bool]:
    if not text.startswith(THREADED_COMMENT_MARKER):
        return text.strip(), False
    marker = "\nComment:\n"
    if marker in text:
        text = text.split(marker, 1)[1]
    return "\n".join(line.strip() for line in text.strip().splitlines()).strip(), True


def header_candidates(
    worksheet: Any,
    cells: list[Any],
    bounds: tuple[int, int, int, int],
) -> dict[int, list[str]]:
    """Find conservative header cues without trying to infer a full table model."""

    from openpyxl.utils.cell import range_boundaries

    first_row, last_row, first_column, last_column = bounds
    cells_by_row: dict[int, list[Any]] = defaultdict(list)
    for cell in cells:
        cells_by_row[cell.row].append(cell)

    table_headers: dict[int, list[str]] = defaultdict(list)
    for name in worksheet.tables.keys():
        table_first_column, table_header_row, table_last_column, _ = range_boundaries(
            worksheet.tables[name].ref
        )
        if (
            first_row <= table_header_row <= last_row
            and table_last_column >= first_column
            and table_first_column <= last_column
        ):
            table_headers[table_header_row].append(name)

    merged_rows = set()
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_col > merged_range.min_col:
            merged_rows.update(range(merged_range.min_row, merged_range.max_row + 1))

    first_content_row = min((cell.row for cell in cells), default=first_row)
    candidates: dict[int, list[str]] = {}
    for row_number, row_cells in sorted(cells_by_row.items()):
        text_cells = [
            cell
            for cell in sorted(row_cells, key=lambda item: item.column)
            if cell.data_type != "f"
            and isinstance(cell.value, str)
            and cell.value.strip()
        ]
        unique_labels = {cell.value.strip().casefold() for cell in text_cells}
        if len(unique_labels) != len(text_cells):
            continue

        reasons = []
        if row_number in table_headers:
            reasons.append("Excel table " + ", ".join(table_headers[row_number]))

        has_name = "name" in unique_labels and len(text_cells) >= 2
        if has_name and not reasons:
            reasons.append('contains "name"')

        previous_cells = cells_by_row.get(row_number - 1, [])
        starts_section = (
            row_number == first_content_row
            or not previous_cells
            or row_number - 1 in merged_rows
        )
        current_is_merged_heading = row_number in merged_rows
        header_columns = {cell.column for cell in text_cells}
        has_data_below = any(
            len(
                {
                    cell.column
                    for cell in cells_by_row.get(following_row, [])
                    if cell.column in header_columns
                }
            )
            >= min(2, len(header_columns))
            for following_row in range(row_number + 1, min(row_number + 4, last_row + 1))
        )
        if (
            len(text_cells) >= 3
            and starts_section
            and not current_is_merged_heading
            and has_data_below
            and not reasons
        ):
            reasons.append("starts a populated section")

        if reasons:
            candidates[row_number] = reasons
    return candidates


def select_worksheets(workbook: Any, requested: list[str] | None) -> list[Any]:
    if not requested:
        return list(workbook.worksheets)

    selected = []
    seen = set()
    for name in requested:
        if name in seen:
            continue
        if name not in workbook.sheetnames:
            available = ", ".join(repr(title) for title in workbook.sheetnames)
            raise ValueError(f"Worksheet {name!r} not found. Available sheets: {available}.")
        selected.append(workbook[name])
        seen.add(name)
    return selected


def workbook_defined_names(workbook: Any) -> list[str]:
    names = []
    for name, definition in workbook.defined_names.items():
        target = getattr(definition, "attr_text", None)
        names.append(f"{name}={target}" if target else name)
    return names


def inspect_workbook(arguments: argparse.Namespace) -> str:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "This script requires openpyxl. Install it with: python -m pip install openpyxl"
        ) from error

    path = arguments.workbook.resolve()
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Workbook must be an .xlsx or .xlsm file.")
    if arguments.max_cells < 0:
        raise ValueError("--max-cells cannot be negative.")
    if arguments.max_formula_patterns < 0:
        raise ValueError("--max-formula-patterns cannot be negative.")

    keep_vba = path.suffix.lower() == ".xlsm"
    formula_workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
    )
    cached_workbook = load_workbook(
        path,
        read_only=False,
        data_only=True,
        keep_vba=keep_vba,
        keep_links=True,
    )

    try:
        worksheets = select_worksheets(formula_workbook, arguments.sheet)
        if arguments.cell_range and len(worksheets) != 1:
            raise ValueError("--range requires exactly one worksheet selected with --sheet.")

        lines = [
            f"Workbook: {path.name} ({human_size(path.stat().st_size)})",
            "Sheets: "
            + ", ".join(
                f"{json.dumps(sheet.title, ensure_ascii=False)} [{sheet.sheet_state}]"
                for sheet in formula_workbook.worksheets
            ),
            "Formula results are saved Excel caches; this tool does not recalculate or verify freshness.",
        ]
        defined_names = workbook_defined_names(formula_workbook)
        if defined_names:
            lines.append("Defined names: " + "; ".join(defined_names))

        workbook_missing_caches = 0
        for worksheet in worksheets:
            cached_worksheet = cached_workbook[worksheet.title]
            bounds = selected_bounds(worksheet, arguments.cell_range)
            cells = [cell for cell in iter_selected_cells(worksheet, bounds) if cell.value is not None]
            formulas: list[FormulaInfo] = []
            annotations: list[str] = []

            for cell in iter_selected_cells(worksheet, bounds):
                if cell.data_type == "f":
                    cached_value = cached_worksheet.cell(cell.row, cell.column).value
                    formulas.append(
                        FormulaInfo(
                            coordinate=cell.coordinate,
                            column=get_column_letter(cell.column),
                            row=cell.row,
                            expression=str(cell.value),
                            cached_value=cached_value,
                        )
                    )
                if cell.comment is not None:
                    comment, was_threaded = clean_comment(cell.comment.text)
                    kind = "threaded comment" if was_threaded else "comment"
                    author = cell.comment.author
                    author_text = f" by {author}" if author and author != "None" else ""
                    annotations.append(
                        f"- {cell.coordinate} {kind}{author_text}: {format_value(comment)}"
                    )
                if cell.hyperlink is not None:
                    target = cell.hyperlink.target or cell.hyperlink.location
                    annotations.append(
                        f"- {cell.coordinate} hyperlink: {format_value(target)}"
                    )

            cached_count = sum(formula.has_cached_value for formula in formulas)
            missing_count = len(formulas) - cached_count
            workbook_missing_caches += missing_count
            headers = header_candidates(worksheet, cells, bounds)
            scope = arguments.cell_range or "all content"
            lines.extend(
                [
                    "",
                    f"## Sheet {json.dumps(worksheet.title, ensure_ascii=False)}",
                    (
                        f"State: {worksheet.sheet_state} | scope: {scope} | "
                        f"content range: {actual_content_range(cells)} | "
                        f"populated: {len(cells)} | formulas: {len(formulas)} "
                        f"({cached_count} cached, {missing_count} missing)"
                    ),
                ]
            )

            structures = []
            if worksheet.tables:
                structures.append(
                    "tables "
                    + ", ".join(
                        f"{name}={worksheet.tables[name].ref}" for name in worksheet.tables.keys()
                    )
                )
            if worksheet.merged_cells.ranges:
                structures.append(
                    "merged " + ", ".join(str(cell_range) for cell_range in worksheet.merged_cells.ranges)
                )
            if worksheet.auto_filter.ref:
                structures.append(f"filter {worksheet.auto_filter.ref}")
            if worksheet.data_validations.dataValidation:
                structures.append(
                    f"data validations {len(worksheet.data_validations.dataValidation)}"
                )
            hidden = hidden_summary(worksheet)
            if hidden:
                structures.append(f"hidden {hidden}")
            if structures:
                lines.append("Structure: " + "; ".join(structures))

            if headers:
                lines.append(
                    "Header cues: "
                    + "; ".join(
                        f"row {row_number} ({'; '.join(reasons)})"
                        for row_number, reasons in headers.items()
                    )
                )

            formulas_by_coordinate = {formula.coordinate: formula for formula in formulas}
            cells_by_row: dict[int, list[Any]] = defaultdict(list)
            for cell in cells:
                cells_by_row[cell.row].append(cell)

            lines.append("Rows:")
            emitted = 0
            truncated = False
            for row_number in sorted(cells_by_row):
                row_cells = sorted(cells_by_row[row_number], key=lambda cell: cell.column)
                remaining = None if arguments.max_cells == 0 else arguments.max_cells - emitted
                if remaining is not None and remaining <= 0:
                    truncated = True
                    break
                if remaining is not None and len(row_cells) > remaining:
                    if emitted:
                        truncated = True
                        break
                    row_cells = row_cells[:remaining]
                    truncated = True

                rendered_cells = []
                for cell in row_cells:
                    formula = formulas_by_coordinate.get(cell.coordinate)
                    if formula is None:
                        rendered_value = format_value(cell.value)
                    else:
                        rendered_value = format_value(formula.cached_value) + "[f]"
                        if not formula.has_cached_value:
                            rendered_value += "!"
                        if arguments.show_formulas:
                            rendered_value += "{" + formula.expression + "}"
                    rendered_cells.append(
                        f"{get_column_letter(cell.column)}={rendered_value}"
                    )
                header_marker = " [header]" if row_number in headers else ""
                lines.append(
                    f"- {row_number}{header_marker}: " + " | ".join(rendered_cells)
                )
                emitted += len(row_cells)
                if truncated:
                    break

            if not cells:
                lines.append("- (no populated cells)")
            elif truncated or emitted < len(cells):
                omitted = len(cells) - emitted
                lines.append(
                    f"- [truncated: {omitted} populated cell(s) omitted; use --sheet, "
                    "--range, or a larger --max-cells]"
                )

            if formulas:
                patterns: dict[tuple[str, str], list[FormulaInfo]] = defaultdict(list)
                for formula in formulas:
                    normalized = normalize_formula(formula.expression, formula.row)
                    patterns[(formula.column, normalized)].append(formula)
                lines.append("Formula patterns:")
                ordered_patterns = sorted(
                    patterns.items(),
                    key=lambda item: (-len(item[1]), item[1][0].column, item[1][0].row),
                )
                pattern_limit = arguments.max_formula_patterns or len(ordered_patterns)
                for (_, expression), grouped in ordered_patterns[:pattern_limit]:
                    coordinates = contiguous_coordinate_ranges(grouped)
                    lines.append(f"- {coordinates} ({len(grouped)}): {expression}")
                omitted_patterns = len(ordered_patterns) - pattern_limit
                if omitted_patterns > 0:
                    lines.append(
                        f"- [truncated: {omitted_patterns} formula pattern(s) omitted; use a "
                        "focused --range, --show-formulas, or a larger --max-formula-patterns]"
                    )

            if annotations:
                lines.append("Annotations:")
                lines.extend(annotations)

        if workbook_missing_caches:
            lines.extend(
                [
                    "",
                    f"WARNING: {workbook_missing_caches} formula cell(s) have no cached result; "
                    "their displayed null values are not ordinary blank cells.",
                ]
            )
        lines.extend(
            [
                "",
                "Legend: [f] = cached formula result; [f]! = formula without a cached result.",
            ]
        )
        return "\n".join(lines)
    finally:
        formula_workbook.close()
        cached_workbook.close()


def configure_output_encoding() -> None:
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if reconfigure is not None:
            reconfigure(encoding="utf-8", errors="backslashreplace")


def main() -> int:
    configure_output_encoding()
    arguments = parse_arguments()
    try:
        print(inspect_workbook(arguments))
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
