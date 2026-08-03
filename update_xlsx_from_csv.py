"""Sync CSV values into the first worksheet of an Excel change sheet.

Only cells whose column header exists in both files are considered. Rows match by
their ``name`` value. Formula expressions are not overwritten unless
``--overwrite-formulas`` is supplied.

Usage:
    python update_xlsx_from_csv.py balancing.csv changes/changes.xlsx 2
    python update_xlsx_from_csv.py balancing.csv changes/changes.xlsx 2 --columns fire_rate bullet_damage
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Any


PLAIN_NUMBER = re.compile(r"^-?(?:0|[1-9]\d*)(?:\.\d+)?(?:[eE][+-]?\d+)?$")
GROUPED_NUMBER = re.compile(r"^-?(?:0|[1-9]\d{0,2})(?:,\d{3})+(?:\.\d+)?$")


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Sync selected CSV values into the first worksheet of an Excel workbook."
    )
    parser.add_argument("csv", type=Path, help="Authoritative UTF-8 CSV source.")
    parser.add_argument("workbook", type=Path, help="Excel change sheet to update (.xlsx or .xlsm).")
    parser.add_argument(
        "header_row",
        type=int,
        help="1-based row containing headers in the workbook's first worksheet.",
    )
    parser.add_argument(
        "--csv-header-row",
        type=int,
        default=2,
        help="1-based row containing headers in the CSV (default: 2).",
    )
    parser.add_argument(
        "--columns",
        nargs="+",
        help="Only sync these columns; defaults to every shared column except name.",
    )
    parser.add_argument(
        "--overwrite-formulas",
        action="store_true",
        help="Replace formulas in selected cells with CSV values.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report what would change without saving the workbook.",
    )
    return parser.parse_args()


def read_csv_rows(csv_path: Path, header_row: int) -> tuple[list[str], dict[str, list[str]]]:
    if header_row < 1:
        raise ValueError("CSV header row must be at least 1.")

    with csv_path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.reader(source)
        try:
            for _ in range(header_row - 1):
                next(reader)
            headers = next(reader)
        except StopIteration as error:
            raise ValueError(f"CSV has no header row {header_row}: {csv_path}") from error

        if headers.count("name") != 1:
            if "name" not in headers:
                raise ValueError(f"CSV header row {header_row} has no 'name' column.")
            raise ValueError(f"CSV header row {header_row} has duplicate 'name' columns.")
        name_index = headers.index("name")
        rows: dict[str, list[str]] = {}
        row_numbers: dict[str, int] = {}

        for row_number, row in enumerate(reader, start=header_row + 1):
            if len(row) <= name_index or not row[name_index]:
                continue
            name = row[name_index]
            if name in rows:
                raise ValueError(
                    f"{csv_path}: duplicate name {name!r} at rows "
                    f"{row_numbers[name]} and {row_number}."
                )
            row_numbers[name] = row_number
            rows[name] = row

    return headers, rows


def header_positions(values: list[Any]) -> dict[str, list[int]]:
    headers: dict[str, list[int]] = {}
    for index, value in enumerate(values):
        if value is None:
            continue
        name = str(value)
        if name:
            headers.setdefault(name, []).append(index)
    return headers


def unique_headers(positions: dict[str, list[int]]) -> dict[str, int]:
    return {name: indexes[0] for name, indexes in positions.items()}


def is_formula(cell: Any) -> bool:
    return cell.data_type == "f"


def excel_value(value: str) -> Any:
    """Keep CSV identifiers as text while writing ordinary numbers as numbers."""
    if not value:
        return None

    number = value.replace(",", "") if GROUPED_NUMBER.fullmatch(value) else value
    if not PLAIN_NUMBER.fullmatch(number):
        return value
    if "." not in number and "e" not in number.casefold():
        return int(number)
    return float(number)


def sync(arguments: argparse.Namespace) -> tuple[int, int, int, list[str]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError as error:
        raise RuntimeError(
            "This script requires openpyxl. Install it with: python -m pip install openpyxl"
        ) from error

    csv_path = arguments.csv.resolve()
    workbook_path = arguments.workbook.resolve()
    if not csv_path.is_file():
        raise FileNotFoundError(f"CSV not found: {csv_path}")
    if csv_path.suffix.lower() != ".csv":
        raise ValueError("CSV source must be a .csv file.")
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Workbook must be an .xlsx or .xlsm file.")
    if arguments.header_row < 1:
        raise ValueError("Workbook header row must be at least 1.")

    csv_headers, csv_rows = read_csv_rows(csv_path, arguments.csv_header_row)
    csv_positions = header_positions(csv_headers)
    csv_columns = unique_headers(csv_positions)
    workbook = load_workbook(
        workbook_path,
        read_only=False,
        data_only=False,
        keep_vba=workbook_path.suffix.lower() == ".xlsm",
        keep_links=True,
    )

    try:
        worksheet = workbook.worksheets[0]
        first_column, _, last_column, last_row = range_boundaries(
            worksheet.calculate_dimension()
        )

        if arguments.header_row > last_row:
            raise ValueError(
                f"Workbook has no header row {arguments.header_row} in its first worksheet."
            )

        workbook_headers = [
            worksheet.cell(arguments.header_row, column).value
            for column in range(first_column, last_column + 1)
        ]
        workbook_positions = header_positions(workbook_headers)
        workbook_columns = unique_headers(workbook_positions)
        if "name" not in workbook_columns:
            raise ValueError(
                f"Workbook header row {arguments.header_row} has no 'name' column."
            )
        if len(workbook_positions["name"]) > 1:
            raise ValueError(
                f"Workbook header row {arguments.header_row} has duplicate 'name' columns."
            )

        available_columns = (set(csv_columns) & set(workbook_columns)) - {"name"}
        selected_columns = set(arguments.columns) if arguments.columns else available_columns
        unknown_columns = selected_columns - available_columns
        if unknown_columns:
            joined = ", ".join(sorted(unknown_columns))
            raise ValueError(f"Selected column(s) are not shared by both files: {joined}")
        if not selected_columns:
            raise ValueError("No shared columns are available to sync.")

        duplicate_csv_columns = {
            column: csv_positions[column]
            for column in selected_columns
            if len(csv_positions[column]) > 1
        }
        if duplicate_csv_columns:
            details = ", ".join(
                f"{column} (CSV columns {', '.join(str(index + 1) for index in indexes)})"
                for column, indexes in sorted(duplicate_csv_columns.items())
            )
            raise ValueError(f"CSV has duplicate selected header(s): {details}")

        duplicate_workbook_columns = {
            column: workbook_positions[column]
            for column in selected_columns
            if len(workbook_positions[column]) > 1
        }
        if duplicate_workbook_columns:
            details = ", ".join(
                f"{column} (worksheet columns "
                f"{', '.join(str(first_column + index) for index in indexes)})"
                for column, indexes in sorted(duplicate_workbook_columns.items())
            )
            raise ValueError(f"Workbook has duplicate selected header(s): {details}")

        name_column = first_column + workbook_columns["name"]
        matched_rows = 0
        updated_cells = 0
        skipped_formulas = 0
        pending_changes: list[str] = []
        workbook_names: dict[str, int] = {}
        for row_offset in range(arguments.header_row + 1, last_row + 1):
            name = worksheet.cell(row_offset, name_column).value
            if name is None or not str(name):
                continue

            name = str(name)
            if name in workbook_names:
                raise ValueError(
                    f"{workbook_path}: duplicate name {name!r} at rows "
                    f"{workbook_names[name]} and {row_offset}."
                )
            workbook_names[name] = row_offset
            if name not in csv_rows:
                continue

            matched_rows += 1
            csv_row = csv_rows[name]
            for column in sorted(selected_columns):
                column_index = first_column + workbook_columns[column]
                cell = worksheet.cell(row_offset, column_index)
                if is_formula(cell) and not arguments.overwrite_formulas:
                    skipped_formulas += 1
                    continue

                csv_index = csv_columns[column]
                value = excel_value(csv_row[csv_index] if csv_index < len(csv_row) else "")
                if cell.value != value:
                    updated_cells += 1
                    if arguments.dry_run:
                        pending_changes.append(
                            f"Row {row_offset}, name {name!r}, column {column!r}: "
                            f"{cell.value!r} -> {value!r}"
                        )
                    else:
                        cell.value = value

        if not matched_rows:
            raise ValueError("No workbook names matched the CSV source.")
        if not arguments.dry_run and updated_cells:
            workbook.save(workbook_path)
        return matched_rows, updated_cells, skipped_formulas, pending_changes
    finally:
        workbook.close()


def main() -> int:
    arguments = parse_arguments()
    try:
        matched_rows, updated_cells, skipped_formulas, pending_changes = sync(arguments)
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    action = "Would update" if arguments.dry_run else "Updated"
    if arguments.dry_run:
        for change in pending_changes:
            print(f"Would update {change}")
    print(f"{action} {updated_cells} cell(s) across {matched_rows} matched row(s).")
    if skipped_formulas:
        print(f"Skipped {skipped_formulas} formula cell(s); use --overwrite-formulas to replace them.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
