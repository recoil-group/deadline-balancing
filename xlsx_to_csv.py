"""Convert the cached values in an Excel workbook's first sheet to UTF-8 CSV.

Formula cells use the value cached when the workbook was last calculated and
saved. The workbook itself is never modified.

Usage:
    python xlsx_to_csv.py path/to/workbook.xlsx
    python xlsx_to_csv.py path/to/workbook.xlsx path/to/output.csv
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass, field
from pathlib import Path


@dataclass(frozen=True)
class ConversionSummary:
    formula_cells: int
    cached_formula_results: int
    missing_formula_cells: frozenset[tuple[int, int]] = field(default_factory=frozenset)

    @property
    def missing_formula_results(self) -> int:
        return self.formula_cells - self.cached_formula_results


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export the first worksheet of an Excel workbook as a UTF-8 CSV."
    )
    parser.add_argument("input", type=Path, help="Workbook to convert (.xlsx or .xlsm).")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        help="CSV destination (defaults to the input path with a .csv extension).",
    )
    return parser.parse_args()


def csv_value(value: object) -> object:
    if isinstance(value, float):
        return format(value, ".15g")
    return value


def convert(input_path: Path, output_path: Path) -> ConversionSummary:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "This script requires openpyxl. Install it with: python -m pip install openpyxl"
        ) from error

    input_path = input_path.resolve()
    output_path = output_path.resolve()

    if not input_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Input must be an .xlsx or .xlsm workbook.")
    if output_path.suffix.lower() != ".csv":
        raise ValueError("Output must have a .csv extension.")
    if input_path == output_path:
        raise ValueError("Input and output paths must be different.")

    formula_workbook = load_workbook(input_path, read_only=True, data_only=False)
    try:
        formula_cells = {
            (row_index, column_index)
            for row_index, row in enumerate(
                formula_workbook.worksheets[0].iter_rows(), start=1
            )
            for column_index, cell in enumerate(row, start=1)
            if cell.data_type == "f"
        }
    finally:
        formula_workbook.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    cached_formula_results = 0
    cached_formula_positions: set[tuple[int, int]] = set()
    try:
        worksheet = workbook.worksheets[0]
        with open(output_path, "w", newline="", encoding="utf-8-sig") as output_file:
            writer = csv.writer(output_file)
            for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                values = []
                for column_index, cell in enumerate(row, start=1):
                    if (
                        (row_index, column_index) in formula_cells
                        and cell.value is not None
                    ):
                        cached_formula_results += 1
                        cached_formula_positions.add((row_index, column_index))
                    values.append(csv_value(cell.value))
                writer.writerow(values)
    finally:
        workbook.close()

    missing_formula_cells = frozenset(
        position for position in formula_cells
        if position not in cached_formula_positions
    )
    return ConversionSummary(len(formula_cells), cached_formula_results, missing_formula_cells)


def print_formula_summary(summary: ConversionSummary) -> None:
    if not summary.formula_cells:
        return
    message = (
        f"Found {summary.formula_cells} formula cell(s): exported "
        f"{summary.cached_formula_results} cached result(s)"
    )
    if summary.missing_formula_results:
        print(
            f"Warning: {message}; {summary.missing_formula_results} without cached results "
            "were exported as blank.",
            file=sys.stderr,
        )
    else:
        print(f"{message}.")


def main() -> int:
    arguments = parse_arguments()
    output_path = arguments.output or arguments.input.with_suffix(".csv")

    try:
        summary = convert(arguments.input, output_path)
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    print_formula_summary(summary)
    print(f"Exported first worksheet to {output_path.resolve()}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
